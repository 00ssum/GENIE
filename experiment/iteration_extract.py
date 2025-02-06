import re
import os
from datetime import datetime
import pandas as pd
from decimal import Decimal
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import numbers


def parse_time(step, log_lines):
    for line in log_lines:
        columns = line.split()  # 띄어쓰기로 분리
        # step 값이 18번째 열에 위치
        if len(columns) > 18 and columns[17].isdigit() and int(columns[17]) == step:
            timestamp_match = re.search(r"\d{2}/\d{2} \d{2}:\d{2}:\d{2}", line)
            if timestamp_match:
                timestamp = timestamp_match.group()
                return datetime.strptime(timestamp, "%m/%d %H:%M:%S")
    return None


# Validation IID 추출 함수
def parse_validation_iid(step, log_lines):
    relevant_lines = []
    for line in log_lines:
        columns = line.split()  # 띄어쓰기로 분리
        # step 값이 18번째 열 이하인 경우만 수집
        if len(columns) > 18 and columns[17].isdigit() and int(columns[17]) <= step:
            relevant_lines.append(columns)
    
    train_out_values = []
    test_in_values = []
    for columns in relevant_lines:
        try:
            train_out_values.append(Decimal(columns[7]))  # train_out 값 (8번째 열 기준)
            test_in_values.append(Decimal(columns[4]))  # test_in 값 (5번째 열 기준)
        except ValueError:
            continue
    
    # train_out에서 argmax 계산 후 해당하는 test_in 값 반환
    if train_out_values:
        max_index = train_out_values.index(max(train_out_values))
        return test_in_values[max_index]
    return None


dataset = "OfficeHome"
alg = "ERM"
envs = ["[0]", "[1]", "[2]", "[3]"]
files = [
    "250205_19-29-17_resnet50_sgd",
    "250206_01-07-12_resnet50_sgd",
    "250206_06-26-05_resnet50_sgd",
    "250206_11-23-42_resnet50_sgd"]


for env, file in zip(envs, files):
    file_path=os.path.join("/jsm0707/GENIE/train_output/",dataset, alg, env, file)
    log_file_path=os.path.join(file_path, "log.txt" )
    with open(log_file_path, 'r') as f:
        log_data = f.readlines()
    output_file=os.path.join(file_path,"iter.xlsx" )
    steps = [0, 5000, 10000, 15000]
    times = {step: parse_time(step, log_data) for step in steps}

    start_time = times[0] 
    elapsed_times = {step: (times[step] - start_time).total_seconds() if times[step] and start_time else None for step in steps}

    validation_iids = {step: parse_validation_iid(step, log_data) for step in steps}

    results = {
        "Step": steps,
        "Training Time (s)": [elapsed_times[step] for step in steps],
        "Validation IID": [float(validation_iids[step]) for step in steps]
    }

    # DataFrame 생성
    df = pd.DataFrame(results)

    # 엑셀로 저장 (반올림 없이 저장 및 셀 서식 지정)
    df.to_excel(output_file, index=False, engine='openpyxl')

    # 엑셀 파일 서식 설정
    wb = load_workbook(output_file)
    ws = wb.active

    # 'Validation IID' 열에 소수점 서식 지정
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):  # Validation IID는 3번째 열
        for cell in row:
            cell.number_format = "0.000000"  # 소수점 12자리 형식

    wb.save(output_file)

    print(f"결과가 저장되었습니다.")